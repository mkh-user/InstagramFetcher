# region imports
import subprocess as sp
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
# endregion

# region initializing
root = tk.Tk()
# endregion

# region global variables
P_target_username = ""
P_target_list = tk.StringVar()
L_login_mode = tk.StringVar()
L_credentials = []
L_session_id = ""
L_credentials_file = ""
E_all_infos = tk.IntVar()
E_only_usernames = tk.IntVar()
E_format = tk.StringVar()
E_path = ""
E_name = ""
S_delay = -1
S_express_mode = tk.IntVar()
S_parts = -1
O_no_exp_limit = tk.IntVar()
menu_bar = tk.Menu(root)
part_id = ""
# endregion

# region settings
min_row_size = [40,40,40,60,40,40,60,40,40,40,40,40,40,40,40,40,40]
# endregion

# region initializing functions
def set_root(title, width, height, column_count):
	root.title(title)
	window_width = width
	window_height = height
	screen_width = root.winfo_screenwidth()
	screen_height = root.winfo_screenheight()
	position_right = int(screen_width / 2 - window_width / 2)
	position_down = int(screen_height / 2 - window_height / 2)
	root.geometry(f"{window_width}x{window_height}+{position_right}+{position_down}")
	root.minsize(window_width,window_height)
	for column in range(column_count):
		root.columnconfigure(column, minsize=200)
	for row in range(len(min_row_size)):
		root.rowconfigure(row, minsize=min_row_size[row])


def set_manu_bar(menu_dict):
	for menu in menu_dict.keys():
		sub_menu = tk.Menu(menu_bar, tearoff=0)
		for item in menu_dict[menu].keys():
			sub_menu.add_command(label=item, command=menu_dict[menu][item])
		menu_bar.add_cascade(label=menu, menu=sub_menu)
	root.config(menu=menu_bar)
# endregion

# region menu actions
def set_preset():
	messagebox.showerror("COMMAND NOT FOUND!", "This part of application not available.")


def use_preset():
	messagebox.showerror("COMMAND NOT FOUND!", "This part of application not available.")


def docs():
	messagebox.showerror("COMMAND NOT FOUND!", "This part of application not available.")


def about():
	messagebox.showinfo("About Program", """Instagram Fetcher V1.0 Stable
NOTE: Not checked in practice!

License: MIT 2024 By Mahan Khalili""")
# endregion

# region button actions
def check_login_mode():
	match L_login_mode.get():
		case "ssid":
			ssid_text_box.grid(row=3, column=1, padx=10, pady=10)
			ssid_label.grid(row=3, column=0, padx=10, pady=10)
			username_label.grid_forget()
			username_text_box.grid_forget()
			password_label.grid_forget()
			password_text_box.grid_forget()
			xlsx_label.grid_forget()
			xlsx_action_button.grid_forget()
			xlsx_file_path.grid_forget()
		case "username":
			ssid_text_box.grid_forget()
			ssid_label.grid_forget()
			username_label.grid(row=3, column=0, padx=10, pady=10)
			username_text_box.grid(row=3, column=1, padx=10, pady=10)
			password_label.grid(row=3, column=2, padx=10, pady=10)
			password_text_box.grid(row=3, column=3, padx=10, pady=10)
			xlsx_label.grid_forget()
			xlsx_action_button.grid_forget()
			xlsx_file_path.grid_forget()
		case "xlsx":
			ssid_text_box.grid_forget()
			ssid_label.grid_forget()
			username_label.grid_forget()
			username_text_box.grid_forget()
			password_label.grid_forget()
			password_text_box.grid_forget()
			xlsx_label.grid(row=3, column=0, padx=10, pady=10)
			xlsx_action_button.grid(row=3, column=1, padx=10, pady=10)
			xlsx_file_path.grid(row=3, column=2, columnspan=2,  padx=10, pady=10)


def select_source_file():
	global L_credentials_file
	L_credentials_file = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel Workbook", "*.xlsx")])
	if L_credentials_file:
		xlsx_file_path.config(text=f"{L_credentials_file}")
	else:
		messagebox.showerror("ERROR", "Please select a source file")


def select_export_path():
	global E_path
	E_path = filedialog.askdirectory(title="Select a Folder")
	if E_path:
		path_file_path.config(text=f"{E_path}")
	else:
		messagebox.showerror("ERROR", "Please select a folder")
# endregion

def run_export():
	global P_target_username
	global part_id
	for i in range(1):
		command = "python core.py export"

		if check_error(target_username_text_box.get(), "Please enter a username"): break
		P_target_username = target_username_text_box.get()
		command += " -u " + P_target_username

		err = False

		match L_login_mode.get():
			case "ssid":
				if check_error(ssid_text_box.get(), "Please enter a SSID"): break
				command += " -ssid " + ssid_text_box.get()

				run_default_modes(command)


			case "username":
				if check_error(username_text_box.get(), "Please enter a username"): break
				if check_error(password_text_box.get(), "Please enter a password"): break
				command += " -lcrd " + username_text_box.get() + " " + password_text_box.get()

				run_default_modes(command)

			case "xlsx":
				workbook = load_workbook(L_credentials_file)
				sheet = workbook["Sheet1"]
				max_row = sheet.max_row
				for row in range(1, max_row+1):
					if row == 1:
						command += " -t " + P_target_list.get()
					else:
						command += " --part " + part_id

					command += " -lcrd " + sheet.cell(row=row, column=1).value + " " + sheet.cell(row=row, column=2).value

					command = append_other(command)
					if command == "err":
						break

					if max_text_box.get() != "":
						command += " -m " + max_text_box.get()
					else:
						messagebox.showerror("ERROR", "Please enter a max rate")
						err = True
						break

					print(command)

					try:
						result = sp.run(command, shell=True, check=True, text=True, capture_output=True)
						print(f"Output:\n{result.stdout}")
						result_text = result.stdout
						start_index = result_text.find("Part exported under")
						if start_index != -1:
							start_index += len("Part exported under")
							part_id = result_text[start_index:].strip()
						else:
							if row != max_row:
								messagebox.showerror("ERROR", "Part export failed, see console for more information")
								err = True
								break
							else:
								messagebox.showinfo("SUCCESS", "Export successful")

					except sp.CalledProcessError as e:
						print(f"Error:\n{e.stderr}")
						messagebox.showerror("ERROR", "An error occurred, see console for more information")
						err = True
						break
				if err:
					break


def run_default_modes(command):
	command += " -t " + P_target_list.get()

	command = append_other(command)
	if command == "err":
		return

	if max_text_box.get() != "":
		command += " -m " + max_text_box.get()

	print(command)

	try:
		result = sp.run(command, shell=True, check=True, text=True, capture_output=True)
		print(f"Output:\n{result.stdout}")
		messagebox.showinfo("SUCCESS", "Export was successful")

	except sp.CalledProcessError as e:
		print(f"Error:\n{e.stderr}")
		messagebox.showerror("ERROR", "An error occurred, see console for more information")


def append_other(command):
	if E_all_infos.get() == 1:
		command += " --all-infos "

	if E_only_usernames.get() == 1:
		command += " --only-usernames "

	command += " -f " + E_format.get()

	if check_error(E_path, "Please select a folder"): return "err"
	command += " --path \"" + E_path + "\""

	if check_error(file_name_text_box.get(), "Please select a file name"): return "err"
	command += " --name " + file_name_text_box.get()

	if S_express_mode == 1:
		command += " -e "

	elif delay_text_box.get() != "":
		command += " -d " + delay_text_box.get()

	if O_no_exp_limit == 1:
		command += " --no-exp-limit "

	return command

set_root()
set_manu_bar()

label_1 = tk.Label(root, text="Target Username:")
label_1.grid(row=0,column=0, padx=10, pady=10)
def check_error(variant, message):
	if variant == "":
		messagebox.showerror("ERROR", message)
		return True
	return False
# region ui tools
def add_label(text, tooltip, row, column, padx=10, pady=10):
	label = tk.Label(root, text=text)
	label.grid(row=row, column=column, padx=padx, pady=pady)
	ToolTip(label, tooltip)


class ToolTip:
	def __init__(self, widget, text):
		self.widget = widget
		self.text = text
		self.tooltip = None
		self.widget.bind("<Enter>", self.show_tooltip)
		self.widget.bind("<Leave>", self.hide_tooltip)


	def show_tooltip(self, event):
		tooltip_label.config(text=self.text)


	def hide_tooltip(self, event):
		tooltip_label.config(text="")
# endregion

# region start
set_root("Instagram Fetcher V1.1 Alpha", 1024, 540, 20)

menus = {
	"File" : {
		"Set as default preset" : set_preset,
		"Use default preset" : use_preset,
		"Exit" : root.destroy,
	},
	"Help" : {
		"Documentation" : docs,
		"About" : about,
	}
}
set_manu_bar(menus)
# endregion


target_username_text_box = tk.Entry(root)
target_username_text_box.grid(row=0,column=1, padx=10, pady=10)

label_3 = tk.Label(root, text="Target List:")
label_3.grid(row=1, column=0, padx=10, pady=10)

P_target_list.set("followers")
followers_target_list = tk.Radiobutton(root, text="Followers", variable=P_target_list, value="followers", command=check_target_list)
followers_target_list.grid(row=1, column=1, padx=10, pady=10)
following_target_list = tk.Radiobutton(root, text="Following", variable=P_target_list, value="following", command=check_target_list)
following_target_list.grid(row=1, column=2, padx=10, pady=10)
both_target_list = tk.Radiobutton(root, text="Both", variable=P_target_list, value="both", command=check_target_list)
both_target_list.grid(row=1, column=3, padx=10, pady=10)
mutuals_target_list = tk.Radiobutton(root, text="Mutuals", variable=P_target_list, value="mutuals", command=check_target_list)
mutuals_target_list.grid(row=1, column=4, padx=10, pady=10)

label_4 = tk.Label(root, text="Login Mode:")
label_4.grid(row=2, column=0, padx=10, pady=10)

L_login_mode.set("ssid")
ssid_login_mode = tk.Radiobutton(root, text="Session ID", variable=L_login_mode, value="ssid", command=check_login_mode)
ssid_login_mode.grid(row=2, column=1, padx=10, pady=10)
username_login_mode = tk.Radiobutton(root, text="Username Password", variable=L_login_mode, value="username", command=check_login_mode)
username_login_mode.grid(row=2, column=2, padx=10, pady=10)
xlsx_login_mode = tk.Radiobutton(root, text="Use Exel File", variable=L_login_mode, value="xlsx", command=check_login_mode)
xlsx_login_mode.grid(row=2, column=3, padx=10, pady=10)

ssid_label = tk.Label(root, text="Session ID:")
ssid_label.grid(row=3, column=0, padx=10, pady=10)
ssid_text_box = tk.Entry(root)
ssid_text_box.grid(row=3, column=1, padx=10, pady=10)
username_label = tk.Label(root, text="Username:")
username_text_box = tk.Entry(root)
password_label = tk.Label(root, text="Password:")
password_text_box = tk.Entry(root)
xlsx_label = tk.Label(root, text="Excel File:")
xlsx_action_button = tk.Button(root, text="Select File", command=select_source_file)
xlsx_file_path = tk.Label(root, text="")

all_infos_check_box = tk.Checkbutton(root, text="Export All Infos", variable=E_all_infos)
all_infos_check_box.grid(row=8, column=0, padx=10, pady=10)

only_usernames_check_box = tk.Checkbutton(root, text="Only Export Usernames", variable=E_only_usernames)
only_usernames_check_box.grid(row=8, column=1, padx=10, pady=10)

label_5 = tk.Label(root, text="Export Format:")
label_5.grid(row=4, column=0, padx=10, pady=10)

E_format.set("excel")
xlsx_format = tk.Radiobutton(root, text=".xlsx", variable=E_format, value="excel")
xlsx_format.grid(row=4, column=1, padx=10, pady=10)
csv_format = tk.Radiobutton(root, text=".csv", variable=E_format, value="csv")
csv_format.grid(row=4, column=2, padx=10, pady=10)
json_format = tk.Radiobutton(root, text=".json", variable=E_format, value="json")
json_format.grid(row=4, column=3, padx=10, pady=10)

label_6 = tk.Label(root, text="Export Path:")
label_6.grid(row=5, column=0, padx=10, pady=10)

path_action_button = tk.Button(root, text="Select Folder", command=select_export_path)
path_action_button.grid(row=5, column=1, padx=10, pady=10)
path_file_path = tk.Label(root, text="")
path_file_path.grid(row=5, column=2, columnspan=2,  padx=10, pady=10)

label_7 = tk.Label(root, text="File Name:")
label_7.grid(row=6, column=0, padx=10, pady=10)

file_name_text_box = tk.Entry(root)
file_name_text_box.grid(row=6, column=1, padx=10, pady=10)

label_8 = tk.Label(root, text="Delay:")
label_8.grid(row=7, column=0, padx=10, pady=10)

delay_text_box = tk.Entry(root)
delay_text_box.grid(row=7, column=1, padx=10, pady=10)

express_mode_check_box = tk.Checkbutton(root, text="Express Mode", variable=S_express_mode)
express_mode_check_box.grid(row=8, column=2, padx=10, pady=10)

label_9 = tk.Label(root, text="Max Request:")
label_9.grid(row=7, column=2, padx=10, pady=10)

max_text_box = tk.Entry(root)
max_text_box.grid(row=7, column=3, padx=10, pady=10)

no_express_limit_check_box = tk.Checkbutton(root, text="No Express Limit", variable=O_no_exp_limit)
no_express_limit_check_box.grid(row=8, column=3, padx=10, pady=10)

run_button = tk.Button(root, text="Run", command=run_export, height=1, width=20)
run_button.grid(row=9,column=4, padx=10, pady=10)

def add_tooltip():
	ToolTip(label_1, "نام کاربری هدف")
	ToolTip(label_3, "هدف")
	ToolTip(label_4, "روش ورود")
	ToolTip(ssid_login_mode, "استفاده از کد رابط")
	ToolTip(username_login_mode, "تنظیم دستی نام کاربری و رمز عبور")
	ToolTip(xlsx_login_mode, "وارد کردن نام کاربری و رمز عبور از فایل اکسل")
	ToolTip(ssid_label, "کد رابط سرور")
	ToolTip(username_label, "نام کاربری")
	ToolTip(password_label, "رمز عبور")
	ToolTip(xlsx_label, "فایل منبع نام کاربری و رمز عبور ها")
	ToolTip(all_infos_check_box, "استخراج تمام اطلاعات ممکن (بدون ارسال درخواست بیشتر)")
	ToolTip(only_usernames_check_box, "فقط استخراج نام کاربری ها")
	ToolTip(label_5, "فرمت فایل خروجی")
	ToolTip(xlsx_format, "Exel Workbook")
	ToolTip(csv_format, "Comma Separated Values")
	ToolTip(json_format, "Json Format")
	ToolTip(label_6, "مسیر فایل خروجی (پوشه فایل)")
	ToolTip(path_file_path, "مسیر انتخاب شده برای خروجی")
	ToolTip(label_7, "نام فایل خروجی")
	ToolTip(label_8, "تاخیر بین ارسال هر درخواست")
	ToolTip(express_mode_check_box, "حالت سریع (ارسال همزمان تمام درخواست ها) - در تعداد بالا به صورت خودکار غیرفعال می شود")
	ToolTip(no_express_limit_check_box, "فعال نگه داشتن حالت سریع در تعداد بالا (احتمال مسدود شدن)")
	ToolTip(label_9, "حداکثر ارسال درخواست (برای هر نام کاربری)")
	ToolTip(run_button, "آغاز استخراج")

tooltip_label = tk.Label(root)
tooltip_label.grid(row=9, column=0, padx=10, pady=10, columnspan=4)

add_tooltip()

root.mainloop()
